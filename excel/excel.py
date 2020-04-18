#! python3

import openpyxl
import os

#How to Read
os.chdir('[INSERT_PATH]\\spreadsheets')

workbook = openpyxl.load_workbook('example.xlsx')

#To get a particular sheet
sheet = workbook.get_sheet_by_name('Sheet1')

#To get sheet names
workbook.get_sheet_names()

#Will return the data in the format the data exists in Excel.
cell = sheet['A1']
cell.value == sheet['A1'].value == sheet.cell(row = 1,column = 1)

for i in range(1,8):
    print(i, sheet.cell(row=i, column=2).value)

#How to write
wb = openpyxl.Workbook()

wb.get_sheet_by_name('Sheet')
sheet['A1'].value

sheet['A1'] = 42
sheet['A2'] = 'Hello'

wb.save('created_file.xlsx')
sheet2 = wb.create_sheet()
sheet2.title= 'My New Sheet Name'
wb.save('example2.xlsx')
wb.create_sheet(index=0,title='My Other Sheet')
wb.save('example3.xlsx')
