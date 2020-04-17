from openpyxl import Workbook
import os, Datetime

os.chdir('''C:\\Users\\jpeva\\Desktop\\Udemy\\Python_ATBSWP\\excel\\
    spreadsheets''')

wb = Workbook()
ws = wb.active
ws.title = "Report"
ws.sheet_properties.tabColor = "FF0000"

#Convert to Dates
ws['A2':'A10'] = datetime.datetime.now()

#More useful feature is being able to access and make changes across rows.
colC = str(ws['C'])
row10 = ws[10:20]

cols_range = ws['C':'D']

for col in ws.iter_cols(min_row=1, max_col=3, max_row=2, values_only = True):
    for cell in col:
        print(cell)

for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        print(cell)

#If you need to iterate through all the rows or columns of a file, you can instead use the Worksheet.rows/cols

tuple(ws.rows)

tuple(ws.columns) #Not available in read-only

ws['A10'].value = 'Smith'

# Use a formula
ws['A27'] = "=SUM(1,1)"
